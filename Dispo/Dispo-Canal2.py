<<<<<<< HEAD
import sys
import os
import polars as pl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# Variables desde línea de comandos
semana = int(sys.argv[1])  # Primer argumento: Semana (1-4)
tienda = int(sys.argv[2])  # Segundo argumento: Tienda (0 o 1)
faltan = int(sys.argv[3])  # Tercer argumento: Faltan (0 o 1)
ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\Planificación y Compras AFM\Forecast Inbound")

carpetas = [carpeta for carpeta in os.listdir(ruta) if os.path.isdir(os.path.join(ruta, carpeta)) and len(carpeta) >= 7 and carpeta[:4].isdigit() and carpeta[5:7].isdigit()]
carpetas.sort(reverse=True, key=lambda x: datetime.strptime(x[:7], "%Y-%m"))

def obtener_excel_mas_reciente(carpeta_path):
    archivos = [f for f in os.listdir(carpeta_path) if f.lower().endswith('.xlsx')]
    archivos.sort(reverse=True, key=lambda x: os.path.getmtime(os.path.join(carpeta_path, x)))
    return os.path.join(carpeta_path, archivos[0]) if archivos else None

def leer_excel_desde_fila_4(ruta_archivo, hoja):
    df = pl.read_excel(
        ruta_archivo, 
        sheet_name=hoja, 
        read_options={"skip_rows": 1}
    )
    return df

df1 = None
if semana in [1, 2]:
    carpeta_mas_reciente = os.path.join(ruta, carpetas[0])
    print(f"Carpeta más reciente para semana {semana}: {carpeta_mas_reciente}")
    archivo_reciente = obtener_excel_mas_reciente(carpeta_mas_reciente)
    print(f"Archivo más reciente: {archivo_reciente}")
    if archivo_reciente:
        df1 = leer_excel_desde_fila_4(archivo_reciente, hoja="BASE")
elif semana in [3, 4]:
    carpeta_mas_reciente = os.path.join(ruta, carpetas[0])
    print(f"Carpeta más reciente para semana {semana}: {carpeta_mas_reciente}")
    archivo_reciente = obtener_excel_mas_reciente(carpeta_mas_reciente)
    print(f"Archivo más reciente: {archivo_reciente}")
    if archivo_reciente:
        df1 = leer_excel_desde_fila_4(archivo_reciente, hoja="BASE")
    carpeta_penultima = os.path.join(ruta, carpetas[1])
    print(f"Penúltima carpeta: {carpeta_penultima}")
    archivo_penultimo = obtener_excel_mas_reciente(carpeta_penultima)
    print(f"Archivo más reciente de la penúltima carpeta: {archivo_penultimo}")
    if archivo_penultimo:
        df2 = leer_excel_desde_fila_4(archivo_penultimo, hoja="BASE")


# In[1139]:


if 'df1' in locals() and df1 is not None:
    for i, row in enumerate(df1.iter_rows(named=False)):
        if row[0] == "ID":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df1 = df1.slice(i + 1)
            df1 = df1.rename({df1.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break

if 'df2' in locals() and df2 is not None:
    for i, row in enumerate(df2.iter_rows(named=False)):
        if row[0] == "ID":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df2 = df2.slice(i + 1)
            df2 = df2.rename({df2.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break


# In[1140]:


df_final_1 = None
df_final_2 = None

if 'df1' in locals() and df1 is not None:
    columnas_seleccionadas = [0, 1] + list(range(170, 182))
    df_final_1 = df1.select([df1.columns[i] for i in columnas_seleccionadas])

if 'df2' in locals() and df2 is not None:
    columnas_seleccionadas = [0, 1] + list(range(170, 182))
    df_final_2 = df2.select([df2.columns[i] for i in columnas_seleccionadas])


# In[1141]:


df_melt_1 = None
df_melt_2 = None

if 'df_final_1' in locals() and df_final_1 is not None:
    df_melt_1 = df_final_1.unpivot(
        index=df_final_1.columns[:2],
        on=df_final_1.columns[2:]
    )
    df_melt_1 = df_melt_1.with_columns(
        df_melt_1['variable'].str.slice(0, 10).alias('variable')
    )

if 'df_final_2' in locals() and df_final_2 is not None:
    df_melt_2 = df_final_2.unpivot(
        index=df_final_2.columns[:2],
        on=df_final_2.columns[2:]
    )
    df_melt_2 = df_melt_2.with_columns(
        df_melt_2['variable'].str.slice(0, 10).alias('variable')
    )


# In[1142]:


if 'df_melt_1' in locals() and df_melt_1 is not None:
    df_melt_1 = df_melt_1.with_columns(
        pl.col('variable').str.strptime(pl.Date, "%Y-%m-%d").alias('variable')
    )

if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.with_columns(
        pl.col('variable').str.strptime(pl.Date, "%Y-%m-%d").alias('variable')
    )


# In[1143]:


if 'df_melt_1' in locals() and df_melt_1 is not None:
    df_melt_1 = df_melt_1.with_columns(
        pl.col('value').cast(pl.Float64).alias('value')
    ).with_columns(
        (pl.col('value') / 4.33).alias('Fc Semanal')
    )

if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.with_columns(
        pl.col('value').cast(pl.Float64).alias('value')
    ).with_columns(
        (pl.col('value') / 4.33).alias('Fc Semanal')
    )


# In[1144]:


if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.filter(
        pl.col('variable') == df_melt_2.select(pl.col('variable').min())[0, 0]
    )


# In[1145]:


print(df_melt_1[0])


# In[1146]:


print(df_melt_1[0])


# In[1147]:


if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_1 = df_melt_1.vstack(df_melt_2).sort('variable')
print(df_melt_1[0])


# In[1148]:


import os
import polars as pl
from datetime import datetime

ruta_principal = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Tubo Semanal")


carpetas = [
    carpeta for carpeta in os.listdir(ruta_principal)
    if '-' in carpeta and carpeta[:10].replace('-', '').isdigit()
]

fechas_carpetas = [
    (carpeta, datetime.strptime(carpeta[:10], "%Y-%m-%d"))
    for carpeta in carpetas
]

if fechas_carpetas:
    carpeta_reciente = max(fechas_carpetas, key=lambda x: x[1])[0]
    ruta_carpeta_reciente = os.path.join(ruta_principal, carpeta_reciente)

    archivos = [
        archivo for archivo in os.listdir(ruta_carpeta_reciente)
        if archivo.lower().endswith('.xlsx')
    ]

    archivo_stock_s4 = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'stock' in archivo.lower() and 's4' in archivo.lower()), None
    )
    archivo_stock_tiendas = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'stock' in archivo.lower() and 'tiendas' in archivo.lower()), None
    )
    archivo_tr_final_s4 = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'tr final s4 consolidado' in archivo.lower()), None
    )

    df_stock = pl.read_excel(archivo_stock_s4, sheet_name="Sheet1") if archivo_stock_s4 else None
    df_tiendas = pl.read_excel(archivo_stock_tiendas, sheet_name="Sheet1") if archivo_stock_tiendas else None
    df_transito = pl.read_excel(archivo_tr_final_s4, sheet_name="Sheet1") if archivo_tr_final_s4 else None

    print("Carpeta más reciente:", ruta_carpeta_reciente)
    print("Archivo Stock S4:", archivo_stock_s4)
    print("Archivo Stock Tiendas:", archivo_stock_tiendas)
    print("Archivo TR Final S4:", archivo_tr_final_s4)
else:
    print("No se encontraron carpetas válidas.")


# In[1149]:


df_stock_agrupado = df_stock.group_by(['Material', 'Centro', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)
df_stock_agrupado = df_stock_agrupado.with_columns(
    (
        pl.col('Trans./Trasl.') + pl.col('Inspecc.de calidad') + pl.col('Bloqueado')
    ).alias('Control')
)
df_711 = df_stock_agrupado.filter(pl.col('Centro').str.contains('711')).drop('Centro')
df_stock = df_stock_agrupado.filter(~pl.col('Centro').str.contains('711')).drop('Centro')

df_711 = df_711.group_by(['Material', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)

df_stock = df_stock.group_by(['Material', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Control').sum().alias('Control'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)


# In[1150]:


df_711 = df_711.rename({col: 'UE' for col in df_711.columns if 'labon' in col})
print(df_711[0])


# In[1151]:


df_stock = df_stock.rename({col: 'UE' for col in df_stock.columns if 'labon' in col})
print(df_stock[0])


# In[1152]:


df_tiendas_agrupado = df_tiendas.group_by('Ult Eslabon').agg(
    pl.col('Libre utilización').cast(pl.Float64).sum().alias('Suma Libre utilización'),
    pl.col('Trans./Trasl.').cast(pl.Float64).sum().alias('Suma Trans./Trasl.'),
    pl.col('Inspecc.de calidad').cast(pl.Float64).sum().alias('Suma Inspecc.de calidad')
)
stock_tiendas = df_tiendas_agrupado.with_columns(
    (pl.col('Suma Libre utilización') + pl.col('Suma Trans./Trasl.') + pl.col('Suma Inspecc.de calidad')).alias('Stock Tiendas')
)
df_tiendas = stock_tiendas.select(['Ult Eslabon', 'Stock Tiendas'])
df_tiendas = df_tiendas.rename({col: "UE" for col in df_tiendas.columns if "labon" in col.lower()})
print(df_tiendas[0])


# In[1153]:


import os
import polars as pl
from openpyxl import load_workbook

ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Faltantes y Sobrantes Retail\Sobrantes y Faltantes AP - AGP")

archivos = [f for f in os.listdir(ruta) if f.lower().endswith('.xlsx')]

archivos.sort(reverse=True, key=lambda x: os.path.getmtime(os.path.join(ruta, x)))

archivo_reciente = os.path.join(ruta, archivos[0]) if archivos else None

if archivo_reciente:
    print(f"Leyendo el archivo: {archivo_reciente}")
    wb = load_workbook(archivo_reciente, read_only=True)
    hojas = wb.sheetnames
    
    hoja_tb_python = next((hoja for hoja in hojas if 'tb python' in hoja.lower()), None)

    if hoja_tb_python:
        faltantes = pl.read_excel(archivo_reciente, sheet_name=hoja_tb_python)
        print(f"Se leyó la hoja: {hoja_tb_python}")
    else:
        print("No se encontró una hoja que contenga 'tb python'.")
else:
    print("No se encontraron archivos Excel en la ruta.")


# In[1154]:


faltantes = faltantes.rename({col: 'UE' for col in faltantes.columns if 'labon' in col})
print(faltantes[0])


# In[1155]:


df_transito[0]


# In[1156]:


df_transito = df_transito.with_columns(
    pl.when((pl.col("Semana") == 1) & (pl.col("Mes") == 12))
    .then(pl.col("Año") + 1)
    .otherwise(pl.col("Año"))
    .alias("Año")
)
print(df_transito[0])


# In[1157]:


df_transito = df_transito.group_by(['Material', 'Texto breve', 'Año','Semana']).agg(
    pl.col('Cantidad').sum().alias('Cantidad')
)
print(df_transito[0])


# In[1158]:


df_transito = df_transito.with_columns(
    pl.col("Año").cast(pl.Utf8).str.slice(-2, 2).alias("Año")
).with_columns(
    (pl.lit("Sem ") + pl.col("Semana").cast(pl.Utf8) + pl.lit(" (") + pl.col("Año") + pl.lit(")")).alias("Sem")
)


# In[1159]:


import os
import polars as pl
from datetime import datetime
from openpyxl import load_workbook

ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\Planificación y Compras AFM\Open to Buy_OTB")

meses_esp = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}

mes_actual = datetime.now().month
anio_actual = datetime.now().year
mes_siguiente = mes_actual + 1 if mes_actual < 12 else 1
anio_mes = anio_actual if semana in [1, 2] else anio_actual if mes_siguiente > 1 else anio_actual + 1
nombre_mes = meses_esp[mes_actual] if semana in [1, 2] else meses_esp[mes_siguiente]

carpetas = [carpeta for carpeta in os.listdir(ruta) if os.path.isdir(os.path.join(ruta, carpeta)) and len(carpeta) == 4]

carpeta_entrada = next(
    (os.path.join(ruta, carpeta) for carpeta in carpetas if str(anio_mes) in carpeta), None
)

if carpeta_entrada:
    print(f"Ingresando a la carpeta: {carpeta_entrada}")
    carpetas_mes = [carpeta for carpeta in os.listdir(carpeta_entrada) if nombre_mes in carpeta.lower()]
    
    if carpetas_mes:
        carpeta_mes = os.path.join(carpeta_entrada, carpetas_mes[0])
        archivos = [f for f in os.listdir(carpeta_mes) if f.lower().endswith('.xlsx') and 'sp' in f.lower()]
        
        if archivos:
            archivo_reciente = max(archivos, key=lambda x: os.path.getmtime(os.path.join(carpeta_mes, x)))
            archivo_path = os.path.join(carpeta_mes, archivo_reciente)
            
            # Usar openpyxl para obtener los nombres de las hojas
            wb = load_workbook(archivo_path, read_only=True)
            hojas = wb.sheetnames

            # Buscar la hoja que contenga 'base' en el nombre
            hoja_base = next((hoja for hoja in hojas if 'base' in hoja.lower()), None)
            
            if hoja_base:
                df_supply_plan = pl.read_excel(archivo_path, sheet_name=hoja_base)
                print(f"Se leyó el archivo: {archivo_path}, hoja: {hoja_base}")
            else:
                print("No se encontró una hoja llamada 'base'.")
        else:
            print("No se encontró archivo con 'SP' en el nombre.")
    else:
        print(f"No se encontró la carpeta del mes: {nombre_mes}")
else:
    print(f"No se encontró carpeta correspondiente al año {anio_mes}")


# In[1160]:


if 'df_supply_plan' in locals() and df_supply_plan is not None:
    for i, row in enumerate(df_supply_plan.iter_rows(named=False)):
        if row[0] == "Material":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df_supply_plan = df_supply_plan.slice(i + 1)
            df_supply_plan = df_supply_plan.rename({df_supply_plan.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break


# In[1161]:


df1_con_origen = df1.join(
    df_supply_plan.select(["Material", "ORIGEN"]),
    left_on="Ult. Eslabón",
    right_on="Material",
    how="left"
)

ponderacion = df1_con_origen.select([
    "ID", 
    "Ult. Eslabón",
    "Vig", 
    "Canal",
    "VProm",
    "Stock Pond",
    "ORIGEN"
] + df1_con_origen.columns[170:176])

ponderacion = ponderacion.with_columns(
    [pl.col(df1_con_origen.columns[i]).cast(pl.Float64).fill_null(0) for i in range(170, 176)]
)

col_inicio = ponderacion.columns.index("ORIGEN") + 1
columnas_posteriores = ponderacion.columns[col_inicio:]

ponderacion = ponderacion.with_columns(
    pl.when(pl.col("ORIGEN") == "NAC")
    .then(
        sum(pl.col(col) for col in columnas_posteriores[:3]) / len(columnas_posteriores[:3])
    )
    .otherwise(
        sum(pl.col(col) for col in columnas_posteriores[:6]) / len(columnas_posteriores[:6])
    )
    .alias("FC PROM")
)

ponderacion = ponderacion.with_columns(
    (pl.col("FC PROM") / 
     pl.col("FC PROM").sum().over("Ult. Eslabón"))
    .fill_nan(0) 
    .alias("%SKU FC")
)
print(ponderacion.head(10))


# In[1162]:


ponderacion


# In[1163]:


df1.columns[170:176]


# In[1164]:


# # Ruta de guardado
# ruta_guardado_csv = r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\ponderacion_modificado.csv"

# # Guardar como CSV con delimitador personalizado
# ponderacion.write_csv(
#     ruta_guardado_csv,
#     separator=';',  # Delimitador por punto y coma
#     float_precision=2,  # Controlar precisión decimal si es necesario
# )

# # Ajustar el separador decimal a ',' en los valores float
# # Convertimos el archivo para reemplazar los puntos decimales con comas
# with open(ruta_guardado_csv, 'r') as file:
#     contenido = file.read().replace('.', ',')

# # Guardar el contenido ajustado nuevamente
# with open(ruta_guardado_csv, 'w') as file:
#     file.write(contenido)

# print(f"Archivo CSV guardado con delimitador ';' y decimales ',' en: {ruta_guardado_csv}")


# In[1165]:


ponderacion.columns


# In[1166]:


base_dispo = df1.select([
    "ID",
    "Ult. Eslabón",
    "Descripcion Material",
    "Marca",
    "Vig",
    "Canal"   
])
base_dispo = base_dispo.join(
    ponderacion.select(["ID", "%SKU FC"]),
    left_on="ID",
    right_on="ID",
    how="inner"
)
base_dispo = base_dispo.rename({"%SKU FC": "Ponderación por canal"})
column_order = [
    "ID",
    "Ult. Eslabón",
    "Descripcion Material",
    "Marca",
    "Ponderación por canal",
    "Vig",
    "Canal"
]
base_dispo = base_dispo.select(column_order)
base_dispo = base_dispo.with_columns(
   pl.col("Canal").map_elements(lambda x:
       "RETAIL" if x == "CL RETAIL" else
       "MAY" if x in ["CL MAYORISTA", "CL CES 01"] else
       "GT" if x in ["CL EASY", "CL SODIMAC", "CL WALMART", "CL SMU", "CL TOTTUS"] else
       "AGROPLANET" if x == "CL AGROPLANET" else
       "0",
       return_dtype=pl.Utf8
   ).alias("Canal Consolidado")
)
print(base_dispo.select("Canal Consolidado").unique())


# In[1167]:


# Obtener las columnas disponibles en el DataFrame
columnas_disponibles = df_supply_plan.columns

# Lista de columnas que quieres seleccionar
columnas_deseadas = [
    "Sector",
    "Agrupación ClaCom",
    "Categoría ClaCom",
    "SubCategoría ClaCom",
    "Bobinas/Implementos/Otros",
    "COD. PROVEEDOR",
    "PROVEEDOR",
    "ORIGEN",
    "LEAD TIME",
    "Codigos Inquebrables",
    "Comentario Sesión",
    "Comentario Finales",
    "OBS Retail",
    "OBS DERCO"
]

# Identificar columnas no encontradas
columnas_no_encontradas = [col for col in columnas_deseadas if col not in columnas_disponibles]

columnas_no_encontradas


# In[1168]:


renombramientos = {}

columna_segmentacion = [col for col in df_supply_plan.columns if "segmentaci" in col.lower() and "abcd" not in col.lower()]
if columna_segmentacion:
    renombramientos[columna_segmentacion[0]] = "Segmentacion AFM"

columna_forecast = [col for col in df_supply_plan.columns if "6 meses" in col.lower()]
if columna_forecast:
    renombramientos[columna_forecast[0]] = "Prom Forecast 3-6-12"

columna_venta = [col for col in df_supply_plan.columns if "4 - 6 - 12" in col]
if columna_venta:
    renombramientos[columna_venta[0]] = "Prom Venta 3-6-12"

columna_codigos = [col for col in df_supply_plan.columns if "Codigos Inquebrables" in col]
if columna_codigos:
    renombramientos[columna_codigos[0]] = "Codigos Transparentes"

if renombramientos:
    df_supply_plan = df_supply_plan.rename(renombramientos)

columnas_seleccionadas = [
    "Material",
    "Sector",
    "Agrupación ClaCom",
    "SubAgrupación ClaCom",
    "Categoría ClaCom",
    "SubCategoría ClaCom",
    "Bobinas/Implementos/Otros",
    "COD. PROVEEDOR",
    "PROVEEDOR",
    "ORIGEN",
    "LEAD TIME",
    "Codigos Transparentes",
    "Comentario Sesión",
    "Comentario Finales",
    "OBS Retail",
    "OBS DERCO"
]

if "Segmentacion AFM" in renombramientos.values():
    columnas_seleccionadas.append("Segmentacion AFM")
if "Prom Forecast 3-6-12" in renombramientos.values():
    columnas_seleccionadas.append("Prom Forecast 3-6-12")
if "Prom Venta 3-6-12" in renombramientos.values():
    columnas_seleccionadas.append("Prom Venta 3-6-12")

df_supply_plan_filtrado = df_supply_plan.select(columnas_seleccionadas)


# In[1169]:


base_dispo = base_dispo.rename({"Ult. Eslabón": "Material"}).join(
    df_supply_plan_filtrado, 
    on="Material", 
    how="left"
).rename({"Material": "Ult. Eslabón"})


# In[1170]:


base_dispo.head(1)


# In[1171]:


import polars as pl
from datetime import datetime, timedelta
from calendar import monthrange

def semanas_del_mes(fecha):
    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
    _, dias_mes = monthrange(fecha_dt.year, fecha_dt.month)
    inicio_mes = datetime(fecha_dt.year, fecha_dt.month, 1)
    fin_mes = datetime(fecha_dt.year, fecha_dt.month, dias_mes)

    semanas = []
    fecha_actual = inicio_mes
    while fecha_actual <= fin_mes:
        semana_iso = fecha_actual.isocalendar().week
        if semana_iso not in semanas:
            semanas.append(semana_iso)
        fecha_actual += timedelta(days=1)

    return semanas

def expandir_df(df):
    df_pd = df.to_pandas()
    df_expanded = df_pd.assign(Semanas=df_pd["variable"].dt.strftime('%Y-%m-%d').apply(semanas_del_mes)).explode("Semanas")
    return pl.from_pandas(df_expanded)

df_melt_1_expanded = expandir_df(df_melt_1)
fc_sem = df_melt_1_expanded
fc_sem = fc_sem.with_columns(
    pl.col("variable").dt.year().alias("Año")
)
fc_sem = fc_sem.filter(~(
    ((pl.col("Año") == 2024) | (pl.col("Año") == 2025) | (pl.col("Año") == 2026) | (pl.col("Año") == 2027)) &
    (pl.col("Semanas") == 1) &
    (pl.col("variable").dt.month() == 12)
))
fc_sem = fc_sem.with_columns(
    pl.col("Fc Semanal").fill_null(0)
)
fc_sem_agrupado = fc_sem.group_by(
    ["ID", "Ult. Eslabón", "Año", "Semanas"]
).agg(
    pl.col("Fc Semanal").mean().alias("Promedio Fc Semanal")
)

print(fc_sem_agrupado[0])


# In[1172]:


# fc_sem_agrupado['variable'].unique()


# In[1173]:


# valores_unicos = fc_sem.filter(
#     # (pl.col("variable") == pl.datetime(2024, 12, 1)) |
#     (pl.col("variable") == pl.datetime(2024, 12, 1)) 
#     # (pl.col("variable") == pl.datetime(2025, 1, 1))
# ).select("Semanas").unique()

# print(valores_unicos)


# In[1174]:


# fc_sem.to_pandas().to_csv(
#     r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\fc_sem.csv",
#     sep=';',
#     decimal=',',
#     index=False
# )

# fc_sem_agrupado.to_pandas().to_csv(
#     r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\fc_sem_agrupado.csv",
#     sep=';',
#     decimal=',',
#     index=False
# )


# In[1175]:


fc_sem_agrupado.head(1)


# In[1176]:


print(
    fc_sem_agrupado.filter(
        (pl.col("ID") == "1000147CL RETAIL") &
        (pl.col("Semanas") == 48) &
        (pl.col("Año") == 2024)
    )
)


# In[1177]:


fc_sem_agrupado = fc_sem_agrupado.with_columns(
    pl.col("Año").cast(pl.Utf8).str.slice(-2, 2).alias("Año")
).with_columns(
    (pl.lit("Sem ") + pl.col("Semanas").cast(pl.Utf8) + pl.lit(" (") + pl.col("Año") + pl.lit(")")).alias("Sem")
)
print(fc_sem_agrupado[0])


# In[1178]:


fc=fc_sem_agrupado
# Cambiar el nombre de las columnas en 'fc' que contengan 'labon' en su nombre
fc = fc.rename({col: "UE" for col in fc.columns if "labón" in col.lower()})

# Cambiar el nombre de las columnas en 'df_transito' que contengan 'material' en su nombre
df_transito = df_transito.rename({col: "UE" for col in df_transito.columns if "Material" in col.lower()})

# Cambiar el nombre de las columnas en 'base_dispo' que contengan 'labón' en su nombre
base_dispo = base_dispo.rename({col: "UE" for col in base_dispo.columns if "labón" in col.lower()})



# In[1179]:


print(df_711[0])
print(df_stock[0])
print(faltantes[0])


# In[1180]:


print(df_transito[0])
print(fc[0])
print(df_tiendas[0])


# In[1181]:


base_dispo = base_dispo.join(
    df_711.select(["UE", "Libre utilización"]),
    on="UE",
    how="left"
)

base_dispo = base_dispo.rename({"Libre utilización": "711"})
base_dispo = base_dispo.with_columns(
    pl.col("711").fill_null(0)
)
base_dispo = base_dispo.join(
    df_stock.select(["UE", "Libre utilización", "Control"]),
    on="UE",
    how="left"
)

base_dispo = base_dispo.rename({
    "Libre utilización": "Stock CD",
    "Control": "Control"
})
base_dispo = base_dispo.join(
    df_tiendas.select(["UE", "Stock Tiendas"]),
    on="UE",
    how="left"
)
faltantes.columns = [col.strip() for col in faltantes.columns]
base_dispo = base_dispo.join(
    faltantes.select(["UE", "Faltante", "Sobrante", "Stock Objetivo"]),
    on="UE",
    how="left"
)
base_dispo = base_dispo.with_columns([
    pl.col(col).fill_null(0) for col in ["711", "Stock CD", "Control", "Stock Tiendas", "Faltante", "Sobrante", "Stock Objetivo"]
])
print(base_dispo[0])


# In[1182]:


min_semana_2024 = df_transito.filter(pl.col("Año") == "24").select(pl.col("Semana").min()).to_numpy()

print(f"La menor semana del año 2024 es: {min_semana_2024[0][0]}")


# In[1183]:


df_transito = df_transito.sort(by="Semana", descending=False)

print(df_transito)


# In[1184]:


df_transito = df_transito.rename({"Material": "UE"})
df_transito[0]


# In[1185]:


fc.columns


# In[1186]:


fc_pvt = fc.pivot(
    values="Promedio Fc Semanal",
    index=["ID"],
    on="Sem"
)


# In[1187]:


fc_pvt = fc_pvt.fill_null(0)
fc_pvt = fc_pvt.rename({col: f"{col}_fc" for col in fc_pvt.columns if col.startswith("Sem")})
fc_pvt[0]


# In[1188]:


import polars as pl

# Agrupar el DataFrame por 'UE' y 'Sem', y calcular la suma de 'Cantidad'
df_transito_agrupado = (
    df_transito.group_by(['UE', 'Sem'])
    .agg(pl.col('Cantidad').sum().alias('Cantidad'))
)

# Mostrar el resultado
print(df_transito_agrupado)
df_transito=df_transito_agrupado


# In[1189]:


df_transito=df_transito_agrupado


# In[1190]:


df_transito_pivot = df_transito.pivot(
    values="Cantidad",
    index=["UE"],
    on="Sem"
)
transito_pvt = df_transito_pivot


# In[1191]:


transito_pvt = transito_pvt.fill_null(0)
transito_pvt = transito_pvt.rename({col: f"{col}_tr" for col in transito_pvt.columns if col.startswith("Sem")})
transito_pvt[0]


# In[1192]:


import datetime

fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]

print(f"La semana actual es: {semana_actual}")

fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]
año_actual = str(fecha_actual.year)[-2:]

columna_nombre = f"Sem {semana_actual} ({año_actual})"

base_dispo = base_dispo.with_columns(
    pl.when(pl.col("Canal Consolidado") == "RETAIL")
    .then(
        pl.col("Stock CD") * pl.col("Ponderación por canal") - 
        (pl.col("Faltante") if faltan == 1 else pl.lit(0)) +
        (pl.col("Stock Tiendas") if tienda == 1 else pl.lit(0))
    )
    .otherwise(pl.col("Stock CD") * pl.col("Ponderación por canal"))
    .alias(columna_nombre)
)

print(base_dispo)


# In[1243]:


dix=base_dispo


# In[1242]:


base_dispo=dix


# In[1244]:


base_dispo = base_dispo.with_columns(
    [
        pl.col("Control").cast(float),
        pl.col("Ponderación por canal").cast(float),
        pl.col(columna_nombre).cast(float),
    ]
)


# In[1245]:


# Obtener el nombre de la nueva columna: 'Sem (semana actual + 1) (yy)'
columna_nombre_siguiente = f"Sem {semana_actual + 1} ({año_actual})"
columna_nombre_siguiente
columna_tr = columna_nombre_siguiente + "_tr"
columna_fc = columna_nombre_siguiente + "_fc"

if columna_tr in transito_pvt.columns:
    base_dispo = base_dispo.join(
        transito_pvt.filter(
            pl.col("UE").is_in(base_dispo["UE"])
        ).select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
        on="UE",
        how="left"
    )
else:
    print(f"Columna {columna_tr} no encontrada en transito_pvt.")


# In[1246]:


if columna_fc in fc_pvt.columns:
    base_dispo = base_dispo.join(
        fc_pvt.filter(
            pl.col("ID").is_in(base_dispo["ID"])
        ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
        on="ID",
        how="left"
    )
else:
    print(f"Columna {columna_fc} no encontrada en fc_pvt.")

base_dispo = base_dispo.with_columns(
    (
        pl.when(
            (pl.col(columna_nombre) +
             (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
             (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) > 0
        )
        .then(
            pl.when(
                ((pl.col(columna_nombre) +
                  (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
                  (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) -
                 pl.col(columna_fc).fill_null(0)) >= 0
            )
            .then(
                (pl.col(columna_nombre) +
                 (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
                 (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) -
                pl.col(columna_fc).fill_null(0)
            )
            .otherwise(0)
        )
        .otherwise(
            pl.col(columna_nombre) +
            (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
            (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))
        )
    ).alias(columna_nombre_siguiente)
)

# Eliminar columnas solo después de todo el cálculo
base_dispo = base_dispo.drop([columna_tr, columna_fc])


# In[1265]:


resultado = base_dispo.filter(pl.col("ID") == "116519CL RETAIL").select(
    [col for col in base_dispo.columns if col.startswith("Sem 47") or col.startswith("Sem 48")or col.startswith("Sem 49")or col.startswith("Sem 50")]
)

print(resultado)


# In[1225]:


# # Filtrar los datos donde id sea "116519CL RETAIL"
# fila_filtrada = base_dispo.filter(pl.col("ID") == "116519CL RETAIL")

# # Mostrar el valor de columna_nombre_siguiente
# print(fila_filtrada.select(pl.col(columna_nombre_siguiente)))


# In[1249]:


sem_tres = f"Sem {semana_actual + 2} ({año_actual})"
sem_tres


# In[1250]:


sem_dos = f"Sem {semana_actual + 1} ({año_actual})"
sem_dos


# In[1251]:


columna_tr = sem_tres + "_tr"
columna_fc = sem_tres + "_fc"


# In[1253]:


if columna_tr in transito_pvt.columns:
    base_dispo = base_dispo.join(
        transito_pvt.filter(
            pl.col("UE").is_in(base_dispo["UE"])
        ).select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
        on="UE",
        how="left"
    )
else:
    print(f"Columna {columna_tr} no encontrada en transito_pvt.")


# In[1254]:


if columna_fc in fc_pvt.columns:
    base_dispo = base_dispo.join(
        fc_pvt.filter(
            pl.col("ID").is_in(base_dispo["ID"])
        ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
        on="ID",
        how="left"
    )
else:
    print(f"Columna {columna_fc} no encontrada en fc_pvt.")


# In[1255]:


base_dispo = base_dispo.with_columns(
    pl.lit(None).alias(f"Sem {semana_actual + 2} ({año_actual})")
)

base_dispo = base_dispo.with_columns(
    pl.col(columna_nombre_siguiente).alias(f"Sem {semana_actual + 2} ({año_actual})")
)


# In[1233]:


# # Filtrar los datos donde ID es igual a "116519CL RETAIL"
# fila_filtrada = base_dispo.filter(pl.col("ID") == "116519CL RETAIL")

# # Calcular manualmente los valores intermedios y el resultado
# sem_tres_valor = fila_filtrada.select(pl.col(sem_tres).cast(float).fill_null(0)).item()
# sem_tres_tr_valor = fila_filtrada.select(pl.col(f"{sem_tres}_tr").cast(float).fill_null(0)).item()
# ponderacion_valor = fila_filtrada.select(pl.col("Ponderación por canal").cast(float).fill_null(1)).item()
# col_711_valor = fila_filtrada.select(pl.col("711").cast(float).fill_null(0)).item()

# # Realizar el cálculo manual
# resultado_esperado = (
#     sem_tres_valor +
#     (sem_tres_tr_valor * ponderacion_valor) +
#     (col_711_valor * ponderacion_valor)
# )

# # Imprimir los valores intermedios y el resultado esperado
# print(f"sem_tres: {sem_tres_valor}")
# print(f"{sem_tres}_tr: {sem_tres_tr_valor}")
# print(f"Ponderación por canal: {ponderacion_valor}")
# print(f"711: {col_711_valor}")
# print(f"Resultado esperado: {resultado_esperado}")


# In[1256]:


base_dispo = base_dispo.with_columns(
    (
        pl.col(sem_tres) + 
        (pl.col(f"{sem_tres}_tr").cast(float).fill_null(0) * pl.col("Ponderación por canal").cast(float).fill_null(1)) + 
        (pl.col("711").cast(float).fill_null(0) * pl.col("Ponderación por canal").cast(float).fill_null(1))
    )
    .alias(sem_tres)
)

# Evaluación y ajuste de sem_tres
base_dispo = base_dispo.with_columns(
    pl.when(
        pl.col(sem_tres) > 0  # Verifica si el resultado es mayor a 0
    )
    .then(
        pl.when(
            (pl.col(sem_tres) - pl.col(f"{sem_tres}_fc").fill_null(0)) >= 0  # Si la resta no es negativa
        )
        .then(
            pl.col(sem_tres) - pl.col(f"{sem_tres}_fc").fill_null(0)  # Mantiene el resultado de la resta
        )
        .otherwise(0)  # Si la resta es negativa, asigna 0
    )
    .otherwise(pl.col(sem_tres))  # Si la condición no se cumple, deja solo la suma inicial
)
base_dispo = base_dispo.drop([columna_tr, columna_fc])


# In[1263]:


base_dispo2=base_dispo


# In[1262]:


base_dispo=base_dispo2


# In[1237]:


print(base_dispo[3])


# In[1264]:


import datetime
import polars as pl

# Obtener la semana y año actual
fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]
año_actual = int(str(fecha_actual.year)[-2:])

# Número total de semanas en un año ISO
SEMANAS_POR_AÑO = 52

# Iterar por 50 semanas ajustando el cruce de años
for i in range(3, 51):  # 50 semanas a partir de la actual
    try:
        # Calcular la semana y el año correspondientes
        semana_iteracion = semana_actual + i
        if semana_iteracion > SEMANAS_POR_AÑO:
            semana_iteracion -= SEMANAS_POR_AÑO
            año_iteracion = str(año_actual + 1).zfill(2)
        else:
            año_iteracion = str(año_actual).zfill(2)

        # Nombres de las semanas y columnas
        semana_anterior = (
            f"Sem {SEMANAS_POR_AÑO} ({str(int(año_iteracion) - 1).zfill(2)})"
            if semana_iteracion == 1
            else f"Sem {semana_iteracion - 1} ({año_iteracion})"
        )
        semana_actual_nombre = f"Sem {semana_iteracion} ({año_iteracion})"
        columna_tr = f"{semana_actual_nombre}_tr"
        columna_fc = f"{semana_actual_nombre}_fc"

        # Inicializar columnas y procesar datos
        columnas_usadas = []
        if columna_tr in transito_pvt.columns:
            base_dispo = base_dispo.join(
                transito_pvt.filter(pl.col("UE").is_in(base_dispo["UE"]))
                .select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
                on="UE",
                how="left"
            )
            columnas_usadas.append(columna_tr)
        else:
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_tr)
            )
            columnas_usadas.append(f"{columna_tr} (asignado 0)")

        if columna_fc in fc_pvt.columns:
            base_dispo = base_dispo.join(
                fc_pvt.filter(pl.col("ID").is_in(base_dispo["ID"]))
                .select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
                on="ID",
                how="left"
            )
            columnas_usadas.append(columna_fc)
        else:
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_fc)
            )
            columnas_usadas.append(f"{columna_fc} (asignado 0)")

        # Calcular valores para la semana actual
        suma_inicial = (
            pl.col(semana_anterior).fill_null(0) +
            (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1))
        )
        resultado_final = pl.when(suma_inicial > 0).then(
            pl.when(suma_inicial - pl.col(columna_fc).fill_null(0) >= 0)
            .then(suma_inicial - pl.col(columna_fc).fill_null(0))
            .otherwise(0)
        ).otherwise(suma_inicial)
        base_dispo = base_dispo.with_columns(
            resultado_final.alias(semana_actual_nombre)
        )

        # Eliminar columnas temporales
        base_dispo = base_dispo.drop([columna_tr, columna_fc])

        # Mostrar columnas usadas
        print(f"Iteración {semana_actual_nombre}: Se usaron columnas: {', '.join(columnas_usadas)}")

    except Exception as e:
        print(f"Error en la iteración {i} ({semana_actual_nombre}): {e}")


# In[1239]:


# Iterar sobre columnas que cumplan las condiciones
for columna in base_dispo.columns:
    if columna.startswith("Sem") and "(" in columna and "D" not in columna:
        columna_fc = f"{columna}_fc"  # Nombre de la columna_fc correspondiente
        nueva_columna = f"{columna}D"  # Nombre de la nueva columna

        # Verificar y agregar columna_fc desde `fc_pvt` si existe
        if columna_fc in fc_pvt.columns:
            base_dispo = base_dispo.join(
                fc_pvt.filter(
                    pl.col("ID").is_in(base_dispo["ID"])
                ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
                on="ID",
                how="left"
            )
        else:
            # Crear columna_fc con valor 0 si no existe
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_fc)
            )

        # Crear la nueva columna basada en las condiciones especificadas
        base_dispo = base_dispo.with_columns(
            pl.when(pl.col(columna) < 0)  # Condición inicial: Sem < 0
            .then(0)
            .when((pl.col(columna) == 0) & (pl.col(columna_fc) > 0))  # Segunda condición
            .then(0)
            .when((pl.col(columna) == 0) & ((pl.col(columna_fc) == 0) | pl.col(columna_fc).is_null()))  # Tercera condición
            .then(1)
            .when((pl.col(columna) > 0) & ((pl.col(columna_fc) == 0) | pl.col(columna_fc).is_null()))  # Cuarta condición
            .then(1)
            .when((pl.col(columna) > 0) & (pl.col(columna_fc) > 0))  # Quinta condición
            .then(
                pl.when(pl.col(columna) / pl.col(columna_fc) > 1)
                .then(1)
                .otherwise(pl.col(columna) / pl.col(columna_fc))
            )
            .alias(nueva_columna)
        )

        # Eliminar la columna_fc después de la iteración
        base_dispo = base_dispo.drop(columna_fc)
# Seleccionar columnas que terminan en ")D"
columnas_d = [col for col in base_dispo.columns if col.endswith(")D")]

# Multiplicar los valores de estas columnas por 100
base_dispo = base_dispo.with_columns(
    [pl.col(col) * 100 for col in columnas_d]
)

print(f"Se han multiplicado por 100 las siguientes columnas: {', '.join(columnas_d)}")


# In[1240]:


base_dispo


# In[1241]:


import os
from datetime import datetime
import polars as pl

# Definir la ruta base
ruta_base = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Disponibilidad Futura\Dispo_Canales")

# Obtener el año y semana actual
anio_actual = datetime.now().year
semana_actual = datetime.now().isocalendar()[1]

# Crear la carpeta del año actual si no existe
ruta_anio = os.path.join(ruta_base, str(anio_actual))
os.makedirs(ruta_anio, exist_ok=True)

# Crear la carpeta de la semana actual dentro de la carpeta del año
ruta_semana = os.path.join(ruta_anio, f"Sem {semana_actual}")
os.makedirs(ruta_semana, exist_ok=True)

# Rutas para guardar los archivos
ruta_base_dispo = os.path.join(ruta_semana, "base_dispo.xlsx")
ruta_fc_pvt = os.path.join(ruta_semana, "fc_pvt.xlsx")
ruta_transito_pvt = os.path.join(ruta_semana, "transito_pvt.xlsx")

# Guardar los DataFrames en Excel sin aplicar formato
base_dispo.write_excel(ruta_base_dispo)
fc_pvt.write_excel(ruta_fc_pvt)
transito_pvt.write_excel(ruta_transito_pvt)

# Mensaje de confirmación
print(f"Archivos guardados exitosamente en la carpeta: {ruta_semana}")

=======
import sys
import os
import polars as pl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
sys.stdin.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# Variables desde línea de comandos
semana = int(sys.argv[1])  # Primer argumento: Semana (1-4)
tienda = int(sys.argv[2])  # Segundo argumento: Tienda (0 o 1)
faltan = int(sys.argv[3])  # Tercer argumento: Faltan (0 o 1)
ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\Planificación y Compras AFM\Forecast Inbound")

carpetas = [carpeta for carpeta in os.listdir(ruta) if os.path.isdir(os.path.join(ruta, carpeta)) and len(carpeta) >= 7 and carpeta[:4].isdigit() and carpeta[5:7].isdigit()]
carpetas.sort(reverse=True, key=lambda x: datetime.strptime(x[:7], "%Y-%m"))

def obtener_excel_mas_reciente(carpeta_path):
    archivos = [f for f in os.listdir(carpeta_path) if f.lower().endswith('.xlsx')]
    archivos.sort(reverse=True, key=lambda x: os.path.getmtime(os.path.join(carpeta_path, x)))
    return os.path.join(carpeta_path, archivos[0]) if archivos else None

def leer_excel_desde_fila_4(ruta_archivo, hoja):
    df = pl.read_excel(
        ruta_archivo, 
        sheet_name=hoja, 
        read_options={"skip_rows": 1}
    )
    return df

df1 = None
if semana in [1, 2]:
    carpeta_mas_reciente = os.path.join(ruta, carpetas[0])
    print(f"Carpeta más reciente para semana {semana}: {carpeta_mas_reciente}")
    archivo_reciente = obtener_excel_mas_reciente(carpeta_mas_reciente)
    print(f"Archivo más reciente: {archivo_reciente}")
    if archivo_reciente:
        df1 = leer_excel_desde_fila_4(archivo_reciente, hoja="BASE")
elif semana in [3, 4]:
    carpeta_mas_reciente = os.path.join(ruta, carpetas[0])
    print(f"Carpeta más reciente para semana {semana}: {carpeta_mas_reciente}")
    archivo_reciente = obtener_excel_mas_reciente(carpeta_mas_reciente)
    print(f"Archivo más reciente: {archivo_reciente}")
    if archivo_reciente:
        df1 = leer_excel_desde_fila_4(archivo_reciente, hoja="BASE")
    carpeta_penultima = os.path.join(ruta, carpetas[1])
    print(f"Penúltima carpeta: {carpeta_penultima}")
    archivo_penultimo = obtener_excel_mas_reciente(carpeta_penultima)
    print(f"Archivo más reciente de la penúltima carpeta: {archivo_penultimo}")
    if archivo_penultimo:
        df2 = leer_excel_desde_fila_4(archivo_penultimo, hoja="BASE")


# In[1139]:


if 'df1' in locals() and df1 is not None:
    for i, row in enumerate(df1.iter_rows(named=False)):
        if row[0] == "ID":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df1 = df1.slice(i + 1)
            df1 = df1.rename({df1.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break

if 'df2' in locals() and df2 is not None:
    for i, row in enumerate(df2.iter_rows(named=False)):
        if row[0] == "ID":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df2 = df2.slice(i + 1)
            df2 = df2.rename({df2.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break


# In[1140]:


df_final_1 = None
df_final_2 = None

if 'df1' in locals() and df1 is not None:
    columnas_seleccionadas = [0, 1] + list(range(170, 182))
    df_final_1 = df1.select([df1.columns[i] for i in columnas_seleccionadas])

if 'df2' in locals() and df2 is not None:
    columnas_seleccionadas = [0, 1] + list(range(170, 182))
    df_final_2 = df2.select([df2.columns[i] for i in columnas_seleccionadas])


# In[1141]:


df_melt_1 = None
df_melt_2 = None

if 'df_final_1' in locals() and df_final_1 is not None:
    df_melt_1 = df_final_1.unpivot(
        index=df_final_1.columns[:2],
        on=df_final_1.columns[2:]
    )
    df_melt_1 = df_melt_1.with_columns(
        df_melt_1['variable'].str.slice(0, 10).alias('variable')
    )

if 'df_final_2' in locals() and df_final_2 is not None:
    df_melt_2 = df_final_2.unpivot(
        index=df_final_2.columns[:2],
        on=df_final_2.columns[2:]
    )
    df_melt_2 = df_melt_2.with_columns(
        df_melt_2['variable'].str.slice(0, 10).alias('variable')
    )


# In[1142]:


if 'df_melt_1' in locals() and df_melt_1 is not None:
    df_melt_1 = df_melt_1.with_columns(
        pl.col('variable').str.strptime(pl.Date, "%Y-%m-%d").alias('variable')
    )

if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.with_columns(
        pl.col('variable').str.strptime(pl.Date, "%Y-%m-%d").alias('variable')
    )


# In[1143]:


if 'df_melt_1' in locals() and df_melt_1 is not None:
    df_melt_1 = df_melt_1.with_columns(
        pl.col('value').cast(pl.Float64).alias('value')
    ).with_columns(
        (pl.col('value') / 4.33).alias('Fc Semanal')
    )

if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.with_columns(
        pl.col('value').cast(pl.Float64).alias('value')
    ).with_columns(
        (pl.col('value') / 4.33).alias('Fc Semanal')
    )


# In[1144]:


if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_2 = df_melt_2.filter(
        pl.col('variable') == df_melt_2.select(pl.col('variable').min())[0, 0]
    )


# In[1145]:


print(df_melt_1[0])


# In[1146]:


print(df_melt_1[0])


# In[1147]:


if 'df_melt_2' in locals() and df_melt_2 is not None:
    df_melt_1 = df_melt_1.vstack(df_melt_2).sort('variable')
print(df_melt_1[0])


# In[1148]:


import os
import polars as pl
from datetime import datetime

ruta_principal = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Tubo Semanal")


carpetas = [
    carpeta for carpeta in os.listdir(ruta_principal)
    if '-' in carpeta and carpeta[:10].replace('-', '').isdigit()
]

fechas_carpetas = [
    (carpeta, datetime.strptime(carpeta[:10], "%Y-%m-%d"))
    for carpeta in carpetas
]

if fechas_carpetas:
    carpeta_reciente = max(fechas_carpetas, key=lambda x: x[1])[0]
    ruta_carpeta_reciente = os.path.join(ruta_principal, carpeta_reciente)

    archivos = [
        archivo for archivo in os.listdir(ruta_carpeta_reciente)
        if archivo.lower().endswith('.xlsx')
    ]

    archivo_stock_s4 = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'stock' in archivo.lower() and 's4' in archivo.lower()), None
    )
    archivo_stock_tiendas = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'stock' in archivo.lower() and 'tiendas' in archivo.lower()), None
    )
    archivo_tr_final_s4 = next(
        (os.path.join(ruta_carpeta_reciente, archivo)
         for archivo in archivos if 'tr final s4 consolidado' in archivo.lower()), None
    )

    df_stock = pl.read_excel(archivo_stock_s4, sheet_name="Sheet1") if archivo_stock_s4 else None
    df_tiendas = pl.read_excel(archivo_stock_tiendas, sheet_name="Sheet1") if archivo_stock_tiendas else None
    df_transito = pl.read_excel(archivo_tr_final_s4, sheet_name="Sheet1") if archivo_tr_final_s4 else None

    print("Carpeta más reciente:", ruta_carpeta_reciente)
    print("Archivo Stock S4:", archivo_stock_s4)
    print("Archivo Stock Tiendas:", archivo_stock_tiendas)
    print("Archivo TR Final S4:", archivo_tr_final_s4)
else:
    print("No se encontraron carpetas válidas.")


# In[1149]:


df_stock_agrupado = df_stock.group_by(['Material', 'Centro', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)
df_stock_agrupado = df_stock_agrupado.with_columns(
    (
        pl.col('Trans./Trasl.') + pl.col('Inspecc.de calidad') + pl.col('Bloqueado')
    ).alias('Control')
)
df_711 = df_stock_agrupado.filter(pl.col('Centro').str.contains('711')).drop('Centro')
df_stock = df_stock_agrupado.filter(~pl.col('Centro').str.contains('711')).drop('Centro')

df_711 = df_711.group_by(['Material', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)

df_stock = df_stock.group_by(['Material', 'Texto breve de material', 'Ult. Eslabon']).agg(
    [
        pl.col('Libre utilización').sum().alias('Libre utilización'),
        pl.col('Control').sum().alias('Control'),
        pl.col('Trans./Trasl.').sum().alias('Trans./Trasl.'),
        pl.col('Inspecc.de calidad').sum().alias('Inspecc.de calidad'),
        pl.col('Bloqueado').sum().alias('Bloqueado'),
    ]
)


# In[1150]:


df_711 = df_711.rename({col: 'UE' for col in df_711.columns if 'labon' in col})
print(df_711[0])


# In[1151]:


df_stock = df_stock.rename({col: 'UE' for col in df_stock.columns if 'labon' in col})
print(df_stock[0])


# In[1152]:


df_tiendas_agrupado = df_tiendas.group_by('Ult Eslabon').agg(
    pl.col('Libre utilización').cast(pl.Float64).sum().alias('Suma Libre utilización'),
    pl.col('Trans./Trasl.').cast(pl.Float64).sum().alias('Suma Trans./Trasl.'),
    pl.col('Inspecc.de calidad').cast(pl.Float64).sum().alias('Suma Inspecc.de calidad')
)
stock_tiendas = df_tiendas_agrupado.with_columns(
    (pl.col('Suma Libre utilización') + pl.col('Suma Trans./Trasl.') + pl.col('Suma Inspecc.de calidad')).alias('Stock Tiendas')
)
df_tiendas = stock_tiendas.select(['Ult Eslabon', 'Stock Tiendas'])
df_tiendas = df_tiendas.rename({col: "UE" for col in df_tiendas.columns if "labon" in col.lower()})
print(df_tiendas[0])


# In[1153]:


import os
import polars as pl
from openpyxl import load_workbook

ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Faltantes y Sobrantes Retail\Sobrantes y Faltantes AP - AGP")

archivos = [f for f in os.listdir(ruta) if f.lower().endswith('.xlsx')]

archivos.sort(reverse=True, key=lambda x: os.path.getmtime(os.path.join(ruta, x)))

archivo_reciente = os.path.join(ruta, archivos[0]) if archivos else None

if archivo_reciente:
    print(f"Leyendo el archivo: {archivo_reciente}")
    wb = load_workbook(archivo_reciente, read_only=True)
    hojas = wb.sheetnames
    
    hoja_tb_python = next((hoja for hoja in hojas if 'tb python' in hoja.lower()), None)

    if hoja_tb_python:
        faltantes = pl.read_excel(archivo_reciente, sheet_name=hoja_tb_python)
        print(f"Se leyó la hoja: {hoja_tb_python}")
    else:
        print("No se encontró una hoja que contenga 'tb python'.")
else:
    print("No se encontraron archivos Excel en la ruta.")


# In[1154]:


faltantes = faltantes.rename({col: 'UE' for col in faltantes.columns if 'labon' in col})
print(faltantes[0])


# In[1155]:


df_transito[0]


# In[1156]:


df_transito = df_transito.with_columns(
    pl.when((pl.col("Semana") == 1) & (pl.col("Mes") == 12))
    .then(pl.col("Año") + 1)
    .otherwise(pl.col("Año"))
    .alias("Año")
)
print(df_transito[0])


# In[1157]:


df_transito = df_transito.group_by(['Material', 'Texto breve', 'Año','Semana']).agg(
    pl.col('Cantidad').sum().alias('Cantidad')
)
print(df_transito[0])


# In[1158]:


df_transito = df_transito.with_columns(
    pl.col("Año").cast(pl.Utf8).str.slice(-2, 2).alias("Año")
).with_columns(
    (pl.lit("Sem ") + pl.col("Semana").cast(pl.Utf8) + pl.lit(" (") + pl.col("Año") + pl.lit(")")).alias("Sem")
)


# In[1159]:


import os
import polars as pl
from datetime import datetime
from openpyxl import load_workbook

ruta = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\Planificación y Compras AFM\Open to Buy_OTB")

meses_esp = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
    7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}

mes_actual = datetime.now().month
anio_actual = datetime.now().year
mes_siguiente = mes_actual + 1 if mes_actual < 12 else 1
anio_mes = anio_actual if semana in [1, 2] else anio_actual if mes_siguiente > 1 else anio_actual + 1
nombre_mes = meses_esp[mes_actual] if semana in [1, 2] else meses_esp[mes_siguiente]

carpetas = [carpeta for carpeta in os.listdir(ruta) if os.path.isdir(os.path.join(ruta, carpeta)) and len(carpeta) == 4]

carpeta_entrada = next(
    (os.path.join(ruta, carpeta) for carpeta in carpetas if str(anio_mes) in carpeta), None
)

if carpeta_entrada:
    print(f"Ingresando a la carpeta: {carpeta_entrada}")
    carpetas_mes = [carpeta for carpeta in os.listdir(carpeta_entrada) if nombre_mes in carpeta.lower()]
    
    if carpetas_mes:
        carpeta_mes = os.path.join(carpeta_entrada, carpetas_mes[0])
        archivos = [f for f in os.listdir(carpeta_mes) if f.lower().endswith('.xlsx') and 'sp' in f.lower()]
        
        if archivos:
            archivo_reciente = max(archivos, key=lambda x: os.path.getmtime(os.path.join(carpeta_mes, x)))
            archivo_path = os.path.join(carpeta_mes, archivo_reciente)
            
            # Usar openpyxl para obtener los nombres de las hojas
            wb = load_workbook(archivo_path, read_only=True)
            hojas = wb.sheetnames

            # Buscar la hoja que contenga 'base' en el nombre
            hoja_base = next((hoja for hoja in hojas if 'base' in hoja.lower()), None)
            
            if hoja_base:
                df_supply_plan = pl.read_excel(archivo_path, sheet_name=hoja_base)
                print(f"Se leyó el archivo: {archivo_path}, hoja: {hoja_base}")
            else:
                print("No se encontró una hoja llamada 'base'.")
        else:
            print("No se encontró archivo con 'SP' en el nombre.")
    else:
        print(f"No se encontró la carpeta del mes: {nombre_mes}")
else:
    print(f"No se encontró carpeta correspondiente al año {anio_mes}")


# In[1160]:


if 'df_supply_plan' in locals() and df_supply_plan is not None:
    for i, row in enumerate(df_supply_plan.iter_rows(named=False)):
        if row[0] == "Material":
            nuevos_nombres = [str(col) for col in row]
            seen = {}
            for idx, col in enumerate(nuevos_nombres):
                if col in seen:
                    seen[col] += 1
                    nuevos_nombres[idx] = f"{col}_{seen[col]}"
                else:
                    seen[col] = 0
            df_supply_plan = df_supply_plan.slice(i + 1)
            df_supply_plan = df_supply_plan.rename({df_supply_plan.columns[j]: nuevos_nombres[j] for j in range(len(nuevos_nombres))})
            break


# In[1161]:


df1_con_origen = df1.join(
    df_supply_plan.select(["Material", "ORIGEN"]),
    left_on="Ult. Eslabón",
    right_on="Material",
    how="left"
)

ponderacion = df1_con_origen.select([
    "ID", 
    "Ult. Eslabón",
    "Vig", 
    "Canal",
    "VProm",
    "Stock Pond",
    "ORIGEN"
] + df1_con_origen.columns[170:176])

ponderacion = ponderacion.with_columns(
    [pl.col(df1_con_origen.columns[i]).cast(pl.Float64).fill_null(0) for i in range(170, 176)]
)

col_inicio = ponderacion.columns.index("ORIGEN") + 1
columnas_posteriores = ponderacion.columns[col_inicio:]

ponderacion = ponderacion.with_columns(
    pl.when(pl.col("ORIGEN") == "NAC")
    .then(
        sum(pl.col(col) for col in columnas_posteriores[:3]) / len(columnas_posteriores[:3])
    )
    .otherwise(
        sum(pl.col(col) for col in columnas_posteriores[:6]) / len(columnas_posteriores[:6])
    )
    .alias("FC PROM")
)

ponderacion = ponderacion.with_columns(
    (pl.col("FC PROM") / 
     pl.col("FC PROM").sum().over("Ult. Eslabón"))
    .fill_nan(0) 
    .alias("%SKU FC")
)
print(ponderacion.head(10))


# In[1162]:


ponderacion


# In[1163]:


df1.columns[170:176]


# In[1164]:


# # Ruta de guardado
# ruta_guardado_csv = r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\ponderacion_modificado.csv"

# # Guardar como CSV con delimitador personalizado
# ponderacion.write_csv(
#     ruta_guardado_csv,
#     separator=';',  # Delimitador por punto y coma
#     float_precision=2,  # Controlar precisión decimal si es necesario
# )

# # Ajustar el separador decimal a ',' en los valores float
# # Convertimos el archivo para reemplazar los puntos decimales con comas
# with open(ruta_guardado_csv, 'r') as file:
#     contenido = file.read().replace('.', ',')

# # Guardar el contenido ajustado nuevamente
# with open(ruta_guardado_csv, 'w') as file:
#     file.write(contenido)

# print(f"Archivo CSV guardado con delimitador ';' y decimales ',' en: {ruta_guardado_csv}")


# In[1165]:


ponderacion.columns


# In[1166]:


base_dispo = df1.select([
    "ID",
    "Ult. Eslabón",
    "Descripcion Material",
    "Marca",
    "Vig",
    "Canal"   
])
base_dispo = base_dispo.join(
    ponderacion.select(["ID", "%SKU FC"]),
    left_on="ID",
    right_on="ID",
    how="inner"
)
base_dispo = base_dispo.rename({"%SKU FC": "Ponderación por canal"})
column_order = [
    "ID",
    "Ult. Eslabón",
    "Descripcion Material",
    "Marca",
    "Ponderación por canal",
    "Vig",
    "Canal"
]
base_dispo = base_dispo.select(column_order)
base_dispo = base_dispo.with_columns(
   pl.col("Canal").map_elements(lambda x:
       "RETAIL" if x == "CL RETAIL" else
       "MAY" if x in ["CL MAYORISTA", "CL CES 01"] else
       "GT" if x in ["CL EASY", "CL SODIMAC", "CL WALMART", "CL SMU", "CL TOTTUS"] else
       "AGROPLANET" if x == "CL AGROPLANET" else
       "0",
       return_dtype=pl.Utf8
   ).alias("Canal Consolidado")
)
print(base_dispo.select("Canal Consolidado").unique())


# In[1167]:


# Obtener las columnas disponibles en el DataFrame
columnas_disponibles = df_supply_plan.columns

# Lista de columnas que quieres seleccionar
columnas_deseadas = [
    "Sector",
    "Agrupación ClaCom",
    "Categoría ClaCom",
    "SubCategoría ClaCom",
    "Bobinas/Implementos/Otros",
    "COD. PROVEEDOR",
    "PROVEEDOR",
    "ORIGEN",
    "LEAD TIME",
    "Codigos Inquebrables",
    "Comentario Sesión",
    "Comentario Finales",
    "OBS Retail",
    "OBS DERCO"
]

# Identificar columnas no encontradas
columnas_no_encontradas = [col for col in columnas_deseadas if col not in columnas_disponibles]

columnas_no_encontradas


# In[1168]:


renombramientos = {}

columna_segmentacion = [col for col in df_supply_plan.columns if "segmentaci" in col.lower() and "abcd" not in col.lower()]
if columna_segmentacion:
    renombramientos[columna_segmentacion[0]] = "Segmentacion AFM"

columna_forecast = [col for col in df_supply_plan.columns if "6 meses" in col.lower()]
if columna_forecast:
    renombramientos[columna_forecast[0]] = "Prom Forecast 3-6-12"

columna_venta = [col for col in df_supply_plan.columns if "4 - 6 - 12" in col]
if columna_venta:
    renombramientos[columna_venta[0]] = "Prom Venta 3-6-12"

columna_codigos = [col for col in df_supply_plan.columns if "Codigos Inquebrables" in col]
if columna_codigos:
    renombramientos[columna_codigos[0]] = "Codigos Transparentes"

if renombramientos:
    df_supply_plan = df_supply_plan.rename(renombramientos)

columnas_seleccionadas = [
    "Material",
    "Sector",
    "Agrupación ClaCom",
    "SubAgrupación ClaCom",
    "Categoría ClaCom",
    "SubCategoría ClaCom",
    "Bobinas/Implementos/Otros",
    "COD. PROVEEDOR",
    "PROVEEDOR",
    "ORIGEN",
    "LEAD TIME",
    "Codigos Transparentes",
    "Comentario Sesión",
    "Comentario Finales",
    "OBS Retail",
    "OBS DERCO"
]

if "Segmentacion AFM" in renombramientos.values():
    columnas_seleccionadas.append("Segmentacion AFM")
if "Prom Forecast 3-6-12" in renombramientos.values():
    columnas_seleccionadas.append("Prom Forecast 3-6-12")
if "Prom Venta 3-6-12" in renombramientos.values():
    columnas_seleccionadas.append("Prom Venta 3-6-12")

df_supply_plan_filtrado = df_supply_plan.select(columnas_seleccionadas)


# In[1169]:


base_dispo = base_dispo.rename({"Ult. Eslabón": "Material"}).join(
    df_supply_plan_filtrado, 
    on="Material", 
    how="left"
).rename({"Material": "Ult. Eslabón"})


# In[1170]:


base_dispo.head(1)


# In[1171]:


import polars as pl
from datetime import datetime, timedelta
from calendar import monthrange

def semanas_del_mes(fecha):
    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
    _, dias_mes = monthrange(fecha_dt.year, fecha_dt.month)
    inicio_mes = datetime(fecha_dt.year, fecha_dt.month, 1)
    fin_mes = datetime(fecha_dt.year, fecha_dt.month, dias_mes)

    semanas = []
    fecha_actual = inicio_mes
    while fecha_actual <= fin_mes:
        semana_iso = fecha_actual.isocalendar().week
        if semana_iso not in semanas:
            semanas.append(semana_iso)
        fecha_actual += timedelta(days=1)

    return semanas

def expandir_df(df):
    df_pd = df.to_pandas()
    df_expanded = df_pd.assign(Semanas=df_pd["variable"].dt.strftime('%Y-%m-%d').apply(semanas_del_mes)).explode("Semanas")
    return pl.from_pandas(df_expanded)

df_melt_1_expanded = expandir_df(df_melt_1)
fc_sem = df_melt_1_expanded
fc_sem = fc_sem.with_columns(
    pl.col("variable").dt.year().alias("Año")
)
fc_sem = fc_sem.filter(~(
    ((pl.col("Año") == 2024) | (pl.col("Año") == 2025) | (pl.col("Año") == 2026) | (pl.col("Año") == 2027)) &
    (pl.col("Semanas") == 1) &
    (pl.col("variable").dt.month() == 12)
))
fc_sem = fc_sem.with_columns(
    pl.col("Fc Semanal").fill_null(0)
)
fc_sem_agrupado = fc_sem.group_by(
    ["ID", "Ult. Eslabón", "Año", "Semanas"]
).agg(
    pl.col("Fc Semanal").mean().alias("Promedio Fc Semanal")
)

print(fc_sem_agrupado[0])


# In[1172]:


# fc_sem_agrupado['variable'].unique()


# In[1173]:


# valores_unicos = fc_sem.filter(
#     # (pl.col("variable") == pl.datetime(2024, 12, 1)) |
#     (pl.col("variable") == pl.datetime(2024, 12, 1)) 
#     # (pl.col("variable") == pl.datetime(2025, 1, 1))
# ).select("Semanas").unique()

# print(valores_unicos)


# In[1174]:


# fc_sem.to_pandas().to_csv(
#     r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\fc_sem.csv",
#     sep=';',
#     decimal=',',
#     index=False
# )

# fc_sem_agrupado.to_pandas().to_csv(
#     r"C:\Users\etorres.DERCOPARTS\OneDrive - DERCO CHILE REPUESTOS SpA\Downloads\fc_sem_agrupado.csv",
#     sep=';',
#     decimal=',',
#     index=False
# )


# In[1175]:


fc_sem_agrupado.head(1)


# In[1176]:


print(
    fc_sem_agrupado.filter(
        (pl.col("ID") == "1000147CL RETAIL") &
        (pl.col("Semanas") == 48) &
        (pl.col("Año") == 2024)
    )
)


# In[1177]:


fc_sem_agrupado = fc_sem_agrupado.with_columns(
    pl.col("Año").cast(pl.Utf8).str.slice(-2, 2).alias("Año")
).with_columns(
    (pl.lit("Sem ") + pl.col("Semanas").cast(pl.Utf8) + pl.lit(" (") + pl.col("Año") + pl.lit(")")).alias("Sem")
)
print(fc_sem_agrupado[0])


# In[1178]:


fc=fc_sem_agrupado
# Cambiar el nombre de las columnas en 'fc' que contengan 'labon' en su nombre
fc = fc.rename({col: "UE" for col in fc.columns if "labón" in col.lower()})

# Cambiar el nombre de las columnas en 'df_transito' que contengan 'material' en su nombre
df_transito = df_transito.rename({col: "UE" for col in df_transito.columns if "Material" in col.lower()})

# Cambiar el nombre de las columnas en 'base_dispo' que contengan 'labón' en su nombre
base_dispo = base_dispo.rename({col: "UE" for col in base_dispo.columns if "labón" in col.lower()})



# In[1179]:


print(df_711[0])
print(df_stock[0])
print(faltantes[0])


# In[1180]:


print(df_transito[0])
print(fc[0])
print(df_tiendas[0])


# In[1181]:


base_dispo = base_dispo.join(
    df_711.select(["UE", "Libre utilización"]),
    on="UE",
    how="left"
)

base_dispo = base_dispo.rename({"Libre utilización": "711"})
base_dispo = base_dispo.with_columns(
    pl.col("711").fill_null(0)
)
base_dispo = base_dispo.join(
    df_stock.select(["UE", "Libre utilización", "Control"]),
    on="UE",
    how="left"
)

base_dispo = base_dispo.rename({
    "Libre utilización": "Stock CD",
    "Control": "Control"
})
base_dispo = base_dispo.join(
    df_tiendas.select(["UE", "Stock Tiendas"]),
    on="UE",
    how="left"
)
faltantes.columns = [col.strip() for col in faltantes.columns]
base_dispo = base_dispo.join(
    faltantes.select(["UE", "Faltante", "Sobrante", "Stock Objetivo"]),
    on="UE",
    how="left"
)
base_dispo = base_dispo.with_columns([
    pl.col(col).fill_null(0) for col in ["711", "Stock CD", "Control", "Stock Tiendas", "Faltante", "Sobrante", "Stock Objetivo"]
])
print(base_dispo[0])


# In[1182]:


min_semana_2024 = df_transito.filter(pl.col("Año") == "24").select(pl.col("Semana").min()).to_numpy()

print(f"La menor semana del año 2024 es: {min_semana_2024[0][0]}")


# In[1183]:


df_transito = df_transito.sort(by="Semana", descending=False)

print(df_transito)


# In[1184]:


df_transito = df_transito.rename({"Material": "UE"})
df_transito[0]


# In[1185]:


fc.columns


# In[1186]:


fc_pvt = fc.pivot(
    values="Promedio Fc Semanal",
    index=["ID"],
    on="Sem"
)


# In[1187]:


fc_pvt = fc_pvt.fill_null(0)
fc_pvt = fc_pvt.rename({col: f"{col}_fc" for col in fc_pvt.columns if col.startswith("Sem")})
fc_pvt[0]


# In[1188]:


import polars as pl

# Agrupar el DataFrame por 'UE' y 'Sem', y calcular la suma de 'Cantidad'
df_transito_agrupado = (
    df_transito.group_by(['UE', 'Sem'])
    .agg(pl.col('Cantidad').sum().alias('Cantidad'))
)

# Mostrar el resultado
print(df_transito_agrupado)
df_transito=df_transito_agrupado


# In[1189]:


df_transito=df_transito_agrupado


# In[1190]:


df_transito_pivot = df_transito.pivot(
    values="Cantidad",
    index=["UE"],
    on="Sem"
)
transito_pvt = df_transito_pivot


# In[1191]:


transito_pvt = transito_pvt.fill_null(0)
transito_pvt = transito_pvt.rename({col: f"{col}_tr" for col in transito_pvt.columns if col.startswith("Sem")})
transito_pvt[0]


# In[1192]:


import datetime

fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]

print(f"La semana actual es: {semana_actual}")

fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]
año_actual = str(fecha_actual.year)[-2:]

columna_nombre = f"Sem {semana_actual} ({año_actual})"

base_dispo = base_dispo.with_columns(
    pl.when(pl.col("Canal Consolidado") == "RETAIL")
    .then(
        pl.col("Stock CD") * pl.col("Ponderación por canal") - 
        (pl.col("Faltante") if faltan == 1 else pl.lit(0)) +
        (pl.col("Stock Tiendas") if tienda == 1 else pl.lit(0))
    )
    .otherwise(pl.col("Stock CD") * pl.col("Ponderación por canal"))
    .alias(columna_nombre)
)

print(base_dispo)


# In[1243]:


dix=base_dispo


# In[1242]:


base_dispo=dix


# In[1244]:


base_dispo = base_dispo.with_columns(
    [
        pl.col("Control").cast(float),
        pl.col("Ponderación por canal").cast(float),
        pl.col(columna_nombre).cast(float),
    ]
)


# In[1245]:


# Obtener el nombre de la nueva columna: 'Sem (semana actual + 1) (yy)'
columna_nombre_siguiente = f"Sem {semana_actual + 1} ({año_actual})"
columna_nombre_siguiente
columna_tr = columna_nombre_siguiente + "_tr"
columna_fc = columna_nombre_siguiente + "_fc"

if columna_tr in transito_pvt.columns:
    base_dispo = base_dispo.join(
        transito_pvt.filter(
            pl.col("UE").is_in(base_dispo["UE"])
        ).select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
        on="UE",
        how="left"
    )
else:
    print(f"Columna {columna_tr} no encontrada en transito_pvt.")


# In[1246]:


if columna_fc in fc_pvt.columns:
    base_dispo = base_dispo.join(
        fc_pvt.filter(
            pl.col("ID").is_in(base_dispo["ID"])
        ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
        on="ID",
        how="left"
    )
else:
    print(f"Columna {columna_fc} no encontrada en fc_pvt.")

base_dispo = base_dispo.with_columns(
    (
        pl.when(
            (pl.col(columna_nombre) +
             (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
             (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) > 0
        )
        .then(
            pl.when(
                ((pl.col(columna_nombre) +
                  (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
                  (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) -
                 pl.col(columna_fc).fill_null(0)) >= 0
            )
            .then(
                (pl.col(columna_nombre) +
                 (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
                 (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))) -
                pl.col(columna_fc).fill_null(0)
            )
            .otherwise(0)
        )
        .otherwise(
            pl.col(columna_nombre) +
            (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1)) +
            (pl.col("Control").fill_null(0) * pl.col("Ponderación por canal").fill_null(1))
        )
    ).alias(columna_nombre_siguiente)
)

# Eliminar columnas solo después de todo el cálculo
base_dispo = base_dispo.drop([columna_tr, columna_fc])


# In[1265]:


resultado = base_dispo.filter(pl.col("ID") == "116519CL RETAIL").select(
    [col for col in base_dispo.columns if col.startswith("Sem 47") or col.startswith("Sem 48")or col.startswith("Sem 49")or col.startswith("Sem 50")]
)

print(resultado)


# In[1225]:


# # Filtrar los datos donde id sea "116519CL RETAIL"
# fila_filtrada = base_dispo.filter(pl.col("ID") == "116519CL RETAIL")

# # Mostrar el valor de columna_nombre_siguiente
# print(fila_filtrada.select(pl.col(columna_nombre_siguiente)))


# In[1249]:


sem_tres = f"Sem {semana_actual + 2} ({año_actual})"
sem_tres


# In[1250]:


sem_dos = f"Sem {semana_actual + 1} ({año_actual})"
sem_dos


# In[1251]:


columna_tr = sem_tres + "_tr"
columna_fc = sem_tres + "_fc"


# In[1253]:


if columna_tr in transito_pvt.columns:
    base_dispo = base_dispo.join(
        transito_pvt.filter(
            pl.col("UE").is_in(base_dispo["UE"])
        ).select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
        on="UE",
        how="left"
    )
else:
    print(f"Columna {columna_tr} no encontrada en transito_pvt.")


# In[1254]:


if columna_fc in fc_pvt.columns:
    base_dispo = base_dispo.join(
        fc_pvt.filter(
            pl.col("ID").is_in(base_dispo["ID"])
        ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
        on="ID",
        how="left"
    )
else:
    print(f"Columna {columna_fc} no encontrada en fc_pvt.")


# In[1255]:


base_dispo = base_dispo.with_columns(
    pl.lit(None).alias(f"Sem {semana_actual + 2} ({año_actual})")
)

base_dispo = base_dispo.with_columns(
    pl.col(columna_nombre_siguiente).alias(f"Sem {semana_actual + 2} ({año_actual})")
)


# In[1233]:


# # Filtrar los datos donde ID es igual a "116519CL RETAIL"
# fila_filtrada = base_dispo.filter(pl.col("ID") == "116519CL RETAIL")

# # Calcular manualmente los valores intermedios y el resultado
# sem_tres_valor = fila_filtrada.select(pl.col(sem_tres).cast(float).fill_null(0)).item()
# sem_tres_tr_valor = fila_filtrada.select(pl.col(f"{sem_tres}_tr").cast(float).fill_null(0)).item()
# ponderacion_valor = fila_filtrada.select(pl.col("Ponderación por canal").cast(float).fill_null(1)).item()
# col_711_valor = fila_filtrada.select(pl.col("711").cast(float).fill_null(0)).item()

# # Realizar el cálculo manual
# resultado_esperado = (
#     sem_tres_valor +
#     (sem_tres_tr_valor * ponderacion_valor) +
#     (col_711_valor * ponderacion_valor)
# )

# # Imprimir los valores intermedios y el resultado esperado
# print(f"sem_tres: {sem_tres_valor}")
# print(f"{sem_tres}_tr: {sem_tres_tr_valor}")
# print(f"Ponderación por canal: {ponderacion_valor}")
# print(f"711: {col_711_valor}")
# print(f"Resultado esperado: {resultado_esperado}")


# In[1256]:


base_dispo = base_dispo.with_columns(
    (
        pl.col(sem_tres) + 
        (pl.col(f"{sem_tres}_tr").cast(float).fill_null(0) * pl.col("Ponderación por canal").cast(float).fill_null(1)) + 
        (pl.col("711").cast(float).fill_null(0) * pl.col("Ponderación por canal").cast(float).fill_null(1))
    )
    .alias(sem_tres)
)

# Evaluación y ajuste de sem_tres
base_dispo = base_dispo.with_columns(
    pl.when(
        pl.col(sem_tres) > 0  # Verifica si el resultado es mayor a 0
    )
    .then(
        pl.when(
            (pl.col(sem_tres) - pl.col(f"{sem_tres}_fc").fill_null(0)) >= 0  # Si la resta no es negativa
        )
        .then(
            pl.col(sem_tres) - pl.col(f"{sem_tres}_fc").fill_null(0)  # Mantiene el resultado de la resta
        )
        .otherwise(0)  # Si la resta es negativa, asigna 0
    )
    .otherwise(pl.col(sem_tres))  # Si la condición no se cumple, deja solo la suma inicial
)
base_dispo = base_dispo.drop([columna_tr, columna_fc])


# In[1263]:


base_dispo2=base_dispo


# In[1262]:


base_dispo=base_dispo2


# In[1237]:


print(base_dispo[3])


# In[1264]:


import datetime
import polars as pl

# Obtener la semana y año actual
fecha_actual = datetime.date.today()
semana_actual = fecha_actual.isocalendar()[1]
año_actual = int(str(fecha_actual.year)[-2:])

# Número total de semanas en un año ISO
SEMANAS_POR_AÑO = 52

# Iterar por 50 semanas ajustando el cruce de años
for i in range(3, 51):  # 50 semanas a partir de la actual
    try:
        # Calcular la semana y el año correspondientes
        semana_iteracion = semana_actual + i
        if semana_iteracion > SEMANAS_POR_AÑO:
            semana_iteracion -= SEMANAS_POR_AÑO
            año_iteracion = str(año_actual + 1).zfill(2)
        else:
            año_iteracion = str(año_actual).zfill(2)

        # Nombres de las semanas y columnas
        semana_anterior = (
            f"Sem {SEMANAS_POR_AÑO} ({str(int(año_iteracion) - 1).zfill(2)})"
            if semana_iteracion == 1
            else f"Sem {semana_iteracion - 1} ({año_iteracion})"
        )
        semana_actual_nombre = f"Sem {semana_iteracion} ({año_iteracion})"
        columna_tr = f"{semana_actual_nombre}_tr"
        columna_fc = f"{semana_actual_nombre}_fc"

        # Inicializar columnas y procesar datos
        columnas_usadas = []
        if columna_tr in transito_pvt.columns:
            base_dispo = base_dispo.join(
                transito_pvt.filter(pl.col("UE").is_in(base_dispo["UE"]))
                .select([pl.col("UE"), pl.col(columna_tr).alias(columna_tr)]),
                on="UE",
                how="left"
            )
            columnas_usadas.append(columna_tr)
        else:
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_tr)
            )
            columnas_usadas.append(f"{columna_tr} (asignado 0)")

        if columna_fc in fc_pvt.columns:
            base_dispo = base_dispo.join(
                fc_pvt.filter(pl.col("ID").is_in(base_dispo["ID"]))
                .select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
                on="ID",
                how="left"
            )
            columnas_usadas.append(columna_fc)
        else:
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_fc)
            )
            columnas_usadas.append(f"{columna_fc} (asignado 0)")

        # Calcular valores para la semana actual
        suma_inicial = (
            pl.col(semana_anterior).fill_null(0) +
            (pl.col(columna_tr).fill_null(0) * pl.col("Ponderación por canal").fill_null(1))
        )
        resultado_final = pl.when(suma_inicial > 0).then(
            pl.when(suma_inicial - pl.col(columna_fc).fill_null(0) >= 0)
            .then(suma_inicial - pl.col(columna_fc).fill_null(0))
            .otherwise(0)
        ).otherwise(suma_inicial)
        base_dispo = base_dispo.with_columns(
            resultado_final.alias(semana_actual_nombre)
        )

        # Eliminar columnas temporales
        base_dispo = base_dispo.drop([columna_tr, columna_fc])

        # Mostrar columnas usadas
        print(f"Iteración {semana_actual_nombre}: Se usaron columnas: {', '.join(columnas_usadas)}")

    except Exception as e:
        print(f"Error en la iteración {i} ({semana_actual_nombre}): {e}")


# In[1239]:


# Iterar sobre columnas que cumplan las condiciones
for columna in base_dispo.columns:
    if columna.startswith("Sem") and "(" in columna and "D" not in columna:
        columna_fc = f"{columna}_fc"  # Nombre de la columna_fc correspondiente
        nueva_columna = f"{columna}D"  # Nombre de la nueva columna

        # Verificar y agregar columna_fc desde `fc_pvt` si existe
        if columna_fc in fc_pvt.columns:
            base_dispo = base_dispo.join(
                fc_pvt.filter(
                    pl.col("ID").is_in(base_dispo["ID"])
                ).select([pl.col("ID"), pl.col(columna_fc).alias(columna_fc)]),
                on="ID",
                how="left"
            )
        else:
            # Crear columna_fc con valor 0 si no existe
            base_dispo = base_dispo.with_columns(
                pl.lit(0).alias(columna_fc)
            )

        # Crear la nueva columna basada en las condiciones especificadas
        base_dispo = base_dispo.with_columns(
            pl.when(pl.col(columna) < 0)  # Condición inicial: Sem < 0
            .then(0)
            .when((pl.col(columna) == 0) & (pl.col(columna_fc) > 0))  # Segunda condición
            .then(0)
            .when((pl.col(columna) == 0) & ((pl.col(columna_fc) == 0) | pl.col(columna_fc).is_null()))  # Tercera condición
            .then(1)
            .when((pl.col(columna) > 0) & ((pl.col(columna_fc) == 0) | pl.col(columna_fc).is_null()))  # Cuarta condición
            .then(1)
            .when((pl.col(columna) > 0) & (pl.col(columna_fc) > 0))  # Quinta condición
            .then(
                pl.when(pl.col(columna) / pl.col(columna_fc) > 1)
                .then(1)
                .otherwise(pl.col(columna) / pl.col(columna_fc))
            )
            .alias(nueva_columna)
        )

        # Eliminar la columna_fc después de la iteración
        base_dispo = base_dispo.drop(columna_fc)
# Seleccionar columnas que terminan en ")D"
columnas_d = [col for col in base_dispo.columns if col.endswith(")D")]

# Multiplicar los valores de estas columnas por 100
base_dispo = base_dispo.with_columns(
    [pl.col(col) * 100 for col in columnas_d]
)

print(f"Se han multiplicado por 100 las siguientes columnas: {', '.join(columnas_d)}")


# In[1240]:


base_dispo


# In[1241]:


import os
from datetime import datetime
import polars as pl

# Definir la ruta base
ruta_base = os.path.expanduser(r"~\DERCO CHILE REPUESTOS SpA\Planificación y abastecimiento - Documentos\KPI\Disponibilidad Futura\Dispo_Canales")

# Obtener el año y semana actual
anio_actual = datetime.now().year
semana_actual = datetime.now().isocalendar()[1]

# Crear la carpeta del año actual si no existe
ruta_anio = os.path.join(ruta_base, str(anio_actual))
os.makedirs(ruta_anio, exist_ok=True)

# Crear la carpeta de la semana actual dentro de la carpeta del año
ruta_semana = os.path.join(ruta_anio, f"Sem {semana_actual}")
os.makedirs(ruta_semana, exist_ok=True)

# Rutas para guardar los archivos
ruta_base_dispo = os.path.join(ruta_semana, "base_dispo.xlsx")
ruta_fc_pvt = os.path.join(ruta_semana, "fc_pvt.xlsx")
ruta_transito_pvt = os.path.join(ruta_semana, "transito_pvt.xlsx")

# Guardar los DataFrames en Excel sin aplicar formato
base_dispo.write_excel(ruta_base_dispo)
fc_pvt.write_excel(ruta_fc_pvt)
transito_pvt.write_excel(ruta_transito_pvt)

# Mensaje de confirmación
print(f"Archivos guardados exitosamente en la carpeta: {ruta_semana}")

>>>>>>> ef08e7057f87a951c388a5ea1bd2ded8c841af84
